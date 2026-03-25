"""
╔══════════════════════════════════════════════════════════════╗
║   ERSeP · Backend DDJJ Transporte Interurbano               ║
║   Python 3.10+ · FastAPI · SQLite · Almacenamiento local    ║
╚══════════════════════════════════════════════════════════════╝

Endpoints:
  POST /api/ddjj/enviar           → Empresa envía DDJJ + archivos
  GET  /api/ddjj/mis-ddjj         → Empresa consulta sus propias DDJJ
  GET  /admin/ddjj                → ERSeP lista todas las DDJJ (requiere token)
  GET  /admin/ddjj/{id}/archivo/{doc_id}  → ERSeP descarga un archivo
  GET  /admin/ddjj/export/excel   → ERSeP exporta Excel consolidado
  GET  /admin/estadisticas        → ERSeP ve KPIs del sistema
"""

import os, json, sqlite3, hashlib, shutil
from datetime import datetime
from pathlib import Path
from typing import Optional, List
from contextlib import asynccontextmanager

from fastapi import (FastAPI, UploadFile, File, Form, HTTPException,
                     Depends, Request, status)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── CONFIGURACIÓN ────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
UPLOADS_DIR = Path(os.getenv("UPLOAD_DIR", str(BASE_DIR / "uploads")))
DB_PATH     = Path(os.getenv("DB_PATH", str(BASE_DIR / "ersep_ddjj.db")))
MAX_FILE_MB = 10          # Tamaño máximo por archivo
ADMIN_TOKEN = os.getenv("ERSEP_ADMIN_TOKEN", "ersep-admin-2026-secreto")
ALLOWED_EXTS = {".pdf", ".xlsx", ".xls", ".csv"}

UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH.parent.mkdir(parents=True, exist_ok=True)


# ─── DATABASE SETUP ───────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS ddjj (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cuit TEXT NOT NULL,
        razon_social TEXT NOT NULL,
        anio INTEGER NOT NULL,
        mes INTEGER NOT NULL,
        es_rectificativa INTEGER DEFAULT 0,
        fecha_envio TEXT NOT NULL,
        ip_origen TEXT,
        datos_json TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS archivos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ddjj_id INTEGER NOT NULL REFERENCES ddjj(id),
        doc_id TEXT NOT NULL,
        nombre_original TEXT NOT NULL,
        nombre_disco TEXT NOT NULL,
        tipo_mime TEXT,
        tamaño_bytes INTEGER,
        ruta_relativa TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE INDEX IF NOT EXISTS idx_ddjj_cuit ON ddjj(cuit);
    CREATE INDEX IF NOT EXISTS idx_ddjj_periodo ON ddjj(anio, mes);
    CREATE INDEX IF NOT EXISTS idx_archivos_ddjj ON archivos(ddjj_id);
    """)
    conn.commit()
    conn.close()


# ─── AUTH ─────────────────────────────────────────────────────
security = HTTPBearer(auto_error=False)

def verify_admin(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials or credentials.credentials != ADMIN_TOKEN:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token de administrador inválido.",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return True


# ─── HELPERS ──────────────────────────────────────────────────
def safe_filename(cuit: str, anio: int, mes: int, doc_id: str, original: str) -> str:
    ext = Path(original).suffix.lower()
    safe_cuit = cuit.replace("-", "")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{safe_cuit}_{anio}{str(mes).zfill(2)}_{doc_id}_{ts}{ext}"

def cuit_dir(cuit: str) -> Path:
    d = UPLOADS_DIR / cuit.replace("-", "")
    d.mkdir(exist_ok=True)
    return d

def kpi_calc(datos: dict) -> dict:
    lineas = datos.get("lineas", [])
    km_total = sum(float(l.get("kmT") or 0) * float(l.get("svcs") or 0) for l in lineas)
    ing_total= sum(float(l.get("ing") or 0) for l in lineas)
    choferes = float(datos.get("choferes") or 0)
    adm      = float(datos.get("administrativos") or 0)
    taller   = float(datos.get("taller") or 0)
    empleados= choferes + adm + taller
    vehs     = float(datos.get("vehiculosActivos") or 0)
    lts      = float(datos.get("litrosCombustible") or 0)
    return {
        "km_total":   round(km_total),
        "ing_total":  round(ing_total, 2),
        "empleados":  round(empleados),
        "ing_x_km":   round(ing_total / km_total, 4) if km_total else 0,
        "chs_x_veh":  round(choferes / vehs, 4) if vehs else 0,
        "empl_x_veh": round(empleados / vehs, 4) if vehs else 0,
        "km_x_veh":   round(km_total / vehs) if vehs else 0,
        "lts_x_veh":  round(lts / vehs) if vehs else 0,
        "km_x_lts":   round(km_total / lts, 4) if lts else 0,
    }

MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


# ─── APP ──────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    print("✅ ERSeP Backend iniciado")
    print(f"   DB: {DB_PATH}")
    print(f"   Archivos: {UPLOADS_DIR}")
    print(f"   Admin token: {ADMIN_TOKEN[:8]}...")
    yield

app = FastAPI(
    title="ERSeP · Sistema DDJJ Transporte Interurbano",
    version="2.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ══════════════════════════════════════════════════════════════
#  SERVIR FRONTEND — usa FileResponse para evitar problemas
#  de encoding/truncamiento con archivos grandes
# ══════════════════════════════════════════════════════════════
FRONTEND_FILE = BASE_DIR / "ERSeP_DDJJ_App.html"

@app.get("/", include_in_schema=False)
async def serve_frontend():
    if FRONTEND_FILE.exists():
        return FileResponse(
            path=str(FRONTEND_FILE),
            media_type="text/html; charset=utf-8",
        )
    return HTMLResponse(content="""
    <html><body style='font-family:sans-serif;padding:40px'>
    <h2>⚠️ Frontend no encontrado</h2>
    <p>Colocá el archivo <b>ERSeP_DDJJ_App.html</b> en la misma carpeta que main.py</p>
    <p>Ruta esperada: """ + str(FRONTEND_FILE) + """</p>
    </body></html>
    """, status_code=404)


@app.post("/api/debug", include_in_schema=False)
async def debug_form(request: Request):
    """Endpoint de diagnóstico - muestra qué campos llegan"""
    form = await request.form()
    return {
        "campos_recibidos": list(form.keys()),
        "valores": {k: str(form[k])[:100] for k in form.keys() if not hasattr(form[k], 'filename')},
        "archivos": {k: form[k].filename for k in form.keys() if hasattr(form[k], 'filename')}
    }

@app.get("/api/version", include_in_schema=False)
async def version():
    return {"version": "2.0-fix", "status": "ok", "mensaje": "Backend ERSeP funcionando correctamente"}


# ══════════════════════════════════════════════════════════════
#  ENDPOINTS EMPRESA
# ══════════════════════════════════════════════════════════════

@app.post("/api/ddjj/enviar", summary="Empresa envía DDJJ con archivos")
async def enviar_ddjj(request: Request):
    """
    Recibe la DDJJ completa de una empresa.
    Los archivos son opcionales (se pueden subir de a uno o todos juntos).
    """
    # ── Leer form data manualmente (acepta cualquier encoding de "año") ──
    form = await request.form()

    def fg(key, default=None):
        """Obtener campo del form, tolerante a variantes de nombre"""
        for k in form.keys():
            if k.lower().replace("ñ","n") == key.lower().replace("ñ","n"):
                return form[k]
        return default

    cuit         = fg("cuit")
    razon_social = fg("razon_social")
    mes_raw      = fg("mes")
    datos_json   = fg("datos_json")
    rect_raw     = fg("es_rectificativa", "0")

    # Año: acepta "año", "anio", "año", o cualquier variante unicode corrupta
    anio_raw = None
    for k in form.keys():
        kn = k.lower()
        if kn in ("año","anio","año") or kn.replace("ñ","n") == "anio" or kn.startswith("a") and kn.endswith("o") and len(kn) <= 5:
            anio_raw = form[k]
            break

    # LOG de diagnóstico - imprime en el CMD lo que llega
    print("\n=== FORM RECIBIDO ===")
    print(f"  Campos: {list(form.keys())}")
    print(f"  cuit={cuit} | rs={razon_social} | mes={mes_raw} | anio={anio_raw}")
    print(f"  datos_json len={len(datos_json) if datos_json else 0}")
    print("====================\n")

    # Validar campos obligatorios
    if not cuit or not razon_social or not mes_raw or not datos_json or not anio_raw:
        missing = [k for k,v in [("cuit",cuit),("razon_social",razon_social),("mes",mes_raw),("datos_json",datos_json),("anio",anio_raw)] if not v]
        print(f"❌ FALTANTES: {missing}")
        raise HTTPException(422, f"Campos faltantes: {missing}. Recibidos: {list(form.keys())}")

    try:
        anio = int(anio_raw)
        mes  = int(mes_raw)
        es_rectificativa = int(rect_raw or 0)
    except ValueError as e:
        raise HTTPException(422, f"Error de tipo: {e}")

    # Obtener archivos
    a1 = form.get("a1"); a2 = form.get("a2"); a3 = form.get("a3")
    b1 = form.get("b1"); b2 = form.get("b2")

    # Validar JSON
    try:
        datos = json.loads(datos_json)
    except json.JSONDecodeError:
        raise HTTPException(400, "datos_json inválido")

    # Validar período
    if not (1 <= mes <= 12 and 2020 <= anio <= 2030):
        raise HTTPException(400, "Período inválido")

    ip = request.client.host if request.client else "desconocida"
    fecha_envio = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db()
    try:
        # ── Insertar registro DDJJ ─────────────────────────────
        cur = conn.execute("""
            INSERT INTO ddjj (cuit, razon_social, anio, mes, es_rectificativa,
                              fecha_envio, ip_origen, datos_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (cuit, razon_social, anio, mes, es_rectificativa,
              fecha_envio, ip, datos_json))
        ddjj_id = cur.lastrowid
        conn.commit()

        # ── Guardar archivos ───────────────────────────────────
        archivos_guardados = []
        carpeta = cuit_dir(cuit) / f"{anio}-{str(mes).zfill(2)}"
        carpeta.mkdir(exist_ok=True)

        for doc_id, upload in [("a1",a1),("a2",a2),("a3",a3),("b1",b1),("b2",b2)]:
            if not upload or not upload.filename:
                continue

            # Validar extensión
            ext = Path(upload.filename).suffix.lower()
            if ext not in ALLOWED_EXTS:
                raise HTTPException(400, f"Extensión no permitida: {ext} (archivo {doc_id})")

            # Validar tamaño (leer en chunks)
            contenido = await upload.read()
            if len(contenido) > MAX_FILE_MB * 1024 * 1024:
                raise HTTPException(413, f"Archivo {doc_id} supera {MAX_FILE_MB} MB")

            nombre_disco = safe_filename(cuit, anio, mes, doc_id, upload.filename)
            ruta_abs = carpeta / nombre_disco
            ruta_rel = str(ruta_abs.relative_to(BASE_DIR))

            # Guardar en disco
            with open(ruta_abs, "wb") as f:
                f.write(contenido)

            # Registrar en DB
            conn.execute("""
                INSERT INTO archivos (ddjj_id, doc_id, nombre_original,
                                      nombre_disco, tipo_mime, tamaño_bytes, ruta_relativa)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (ddjj_id, doc_id, upload.filename, nombre_disco,
                  upload.content_type, len(contenido), ruta_rel))
            conn.commit()

            archivos_guardados.append({
                "doc_id": doc_id,
                "nombre": upload.filename,
                "tamaño_kb": round(len(contenido) / 1024, 1),
            })

        return {
            "ok": True,
            "ddjj_id": ddjj_id,
            "mensaje": (
                "✅ Rectificativa recibida correctamente"
                if es_rectificativa else
                "✅ DDJJ recibida y registrada correctamente"
            ),
            "fecha_envio": fecha_envio,
            "archivos_guardados": archivos_guardados,
            "kpi": kpi_calc(datos),
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Error interno: {str(e)}")
    finally:
        conn.close()


@app.get("/api/ddjj/mis-ddjj", summary="Empresa consulta sus propias DDJJ")
def mis_ddjj(cuit: str):
    """Devuelve el historial de presentaciones de una empresa."""
    conn = get_db()
    rows = conn.execute("""
        SELECT d.id, d.anio, d.mes, d.es_rectificativa, d.fecha_envio,
               COUNT(a.id) as cant_archivos
        FROM ddjj d
        LEFT JOIN archivos a ON a.ddjj_id = d.id
        WHERE d.cuit = ?
        GROUP BY d.id
        ORDER BY d.anio DESC, d.mes DESC, d.id DESC
    """, (cuit,)).fetchall()
    conn.close()
    return [
        {
            "id": r["id"],
            "periodo": f"{MESES[r['mes']]} {r['anio']}",
            "anio": r["anio"],
            "mes": r["mes"],
            "es_rectificativa": bool(r["es_rectificativa"]),
            "fecha_envio": r["fecha_envio"],
            "archivos": r["cant_archivos"],
        }
        for r in rows
    ]


# ══════════════════════════════════════════════════════════════
#  ENDPOINTS ADMIN (requieren Bearer token)
# ══════════════════════════════════════════════════════════════

@app.get("/admin/ddjj", summary="Listar todas las DDJJ", dependencies=[Depends(verify_admin)])
def admin_listar(anio: Optional[int] = None, mes: Optional[int] = None, cuit: Optional[str] = None):
    """ERSeP lista presentaciones con filtros opcionales."""
    conn = get_db()
    sql = """
        SELECT d.id, d.cuit, d.razon_social, d.anio, d.mes,
               d.es_rectificativa, d.fecha_envio, d.ip_origen,
               d.datos_json, COUNT(a.id) as cant_archivos
        FROM ddjj d
        LEFT JOIN archivos a ON a.ddjj_id = d.id
        WHERE 1=1
    """
    params = []
    if anio: sql += " AND d.anio = ?"; params.append(anio)
    if mes:  sql += " AND d.mes = ?";  params.append(mes)
    if cuit: sql += " AND d.cuit = ?"; params.append(cuit)
    sql += " GROUP BY d.id ORDER BY d.anio DESC, d.mes DESC, d.id DESC"

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    result = []
    for r in rows:
        datos = json.loads(r["datos_json"])
        result.append({
            "id": r["id"],
            "cuit": r["cuit"],
            "razon_social": r["razon_social"],
            "periodo": f"{MESES[r['mes']]} {r['anio']}",
            "anio": r["anio"],
            "mes": r["mes"],
            "es_rectificativa": bool(r["es_rectificativa"]),
            "fecha_envio": r["fecha_envio"],
            "ip_origen": r["ip_origen"],
            "archivos": r["cant_archivos"],
            "kpi": kpi_calc(datos),
            "datos": datos,
        })
    return result


@app.get("/admin/ddjj/{ddjj_id}", summary="Detalle de una DDJJ", dependencies=[Depends(verify_admin)])
def admin_detalle(ddjj_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM ddjj WHERE id = ?", (ddjj_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "DDJJ no encontrada")

    archivos = conn.execute(
        "SELECT * FROM archivos WHERE ddjj_id = ?", (ddjj_id,)
    ).fetchall()
    conn.close()

    datos = json.loads(row["datos_json"])
    return {
        "id": row["id"],
        "cuit": row["cuit"],
        "razon_social": row["razon_social"],
        "periodo": f"{MESES[row['mes']]} {row['anio']}",
        "fecha_envio": row["fecha_envio"],
        "es_rectificativa": bool(row["es_rectificativa"]),
        "datos": datos,
        "kpi": kpi_calc(datos),
        "archivos": [
            {
                "id": a["id"],
                "doc_id": a["doc_id"],
                "nombre_original": a["nombre_original"],
                "tamaño_kb": round(a["tamaño_bytes"] / 1024, 1),
                "tipo_mime": a["tipo_mime"],
                "url_descarga": f"/admin/archivos/{a['id']}",
            }
            for a in archivos
        ],
    }


@app.get("/admin/archivos-lista/{ddjj_id}", summary="Listar archivos de una DDJJ", dependencies=[Depends(verify_admin)])
def listar_archivos(ddjj_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, doc_id, nombre_original, tamaño_bytes, tipo_mime FROM archivos WHERE ddjj_id = ?",
        (ddjj_id,)
    ).fetchall()
    conn.close()
    return {"archivos": [dict(r) for r in rows]}


@app.get("/admin/archivos/{archivo_id}", summary="Descargar archivo", dependencies=[Depends(verify_admin)])
def descargar_archivo(archivo_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM archivos WHERE id = ?", (archivo_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Archivo no encontrado")

    ruta = BASE_DIR / row["ruta_relativa"]
    if not ruta.exists():
        raise HTTPException(410, "Archivo no disponible en disco")

    return FileResponse(
        path=str(ruta),
        filename=row["nombre_original"],
        media_type=row["tipo_mime"] or "application/octet-stream",
    )


@app.get("/admin/estadisticas", summary="KPIs del sistema", dependencies=[Depends(verify_admin)])
def estadisticas(anio: Optional[int] = None, mes: Optional[int] = None):
    conn = get_db()
    sql = "SELECT datos_json, cuit, razon_social FROM ddjj WHERE 1=1"
    params = []
    if anio: sql += " AND anio = ?"; params.append(anio)
    if mes:  sql += " AND mes = ?";  params.append(mes)

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    empresas = {}
    for r in rows:
        datos = json.loads(r["datos_json"])
        kpi = kpi_calc(datos)
        cuit = r["cuit"]
        if cuit not in empresas or True:  # toma la más reciente
            empresas[cuit] = {"razon_social": r["razon_social"], "kpi": kpi}

    totales = {
        "empresas": len(empresas),
        "km_total": sum(e["kpi"]["km_total"] for e in empresas.values()),
        "ing_total": sum(e["kpi"]["ing_total"] for e in empresas.values()),
        "empleados": sum(e["kpi"]["empleados"] for e in empresas.values()),
    }
    return {"totales": totales, "por_empresa": list(empresas.values())}


@app.get("/admin/ddjj/export/excel", summary="Exportar Excel consolidado", dependencies=[Depends(verify_admin)])
def exportar_excel(anio: Optional[int] = None, mes: Optional[int] = None):
    """Genera un Excel con todas las DDJJ del período seleccionado."""
    conn = get_db()
    sql = "SELECT * FROM ddjj WHERE 1=1"
    params = []
    if anio: sql += " AND anio = ?"; params.append(anio)
    if mes:  sql += " AND mes = ?";  params.append(mes)
    sql += " ORDER BY razon_social"

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"

    navy  = "1F3864"
    blue  = "2E75B6"
    light = "D6E4F0"
    titulo_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=navy)
    sub_fill    = PatternFill("solid", fgColor=blue)
    alt_fill    = PatternFill("solid", fgColor="F4F6FA")

    headers = [
        "CUIT","Razón Social","Período","Estado","Fecha Envío",
        "Km Producidos","Ingresos ($)","Vehículos","Choferes",
        "Empleados Total","Litros","Ingreso/km","Chof/Veh",
        "Empl/Veh","Km/Veh","Lts/Veh","Km/Litro","Archivos"
    ]

    # Título
    ws1.merge_cells("A1:R1")
    ws1["A1"] = f"ERSeP · DDJJ Transporte Interurbano · {MESES[mes] + ' ' + str(anio) if anio and mes else 'Todos los períodos'}"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color=navy)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1.append([])
    ws1.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col)
        cell.font = titulo_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[3].height = 28

    conn2 = get_db()
    for i, r in enumerate(rows):
        datos = json.loads(r["datos_json"])
        kpi = kpi_calc(datos)
        cant_arch = conn2.execute(
            "SELECT COUNT(*) FROM archivos WHERE ddjj_id=?", (r["id"],)
        ).fetchone()[0]

        row_data = [
            r["cuit"], r["razon_social"],
            f"{MESES[r['mes']]} {r['anio']}",
            "Rectificativa" if r["es_rectificativa"] else "Original",
            r["fecha_envio"],
            kpi["km_total"], round(kpi["ing_total"], 2),
            datos.get("vehiculosActivos",""),
            datos.get("choferes",""),
            kpi["empleados"],
            datos.get("litrosCombustible",""),
            kpi["ing_x_km"], kpi["chs_x_veh"], kpi["empl_x_veh"],
            kpi["km_x_veh"], kpi["lts_x_veh"], kpi["km_x_lts"],
            cant_arch,
        ]
        ws1.append(row_data)
        row_num = ws1.max_row
        if i % 2 == 0:
            for col in range(1, len(headers)+1):
                ws1.cell(row=row_num, column=col).fill = alt_fill

    # Anchos
    widths = [18,35,16,14,18,14,14,10,10,14,12,12,10,10,10,10,10,10]
    for col, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Hoja 2: Líneas ────────────────────────────────────────
    ws2 = wb.create_sheet("Líneas por empresa")
    h2 = ["Empresa","CUIT","Corredor","Cód. Línea","Tramo","Km/Trayecto","Servicios","Km Prod.","Ingresos ($)"]
    ws2.append(h2)
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = titulo_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        datos = json.loads(r["datos_json"])
        for l in datos.get("lineas", []):
            km_prod = round(float(l.get("kmT") or 0) * float(l.get("svcs") or 0))
            ws2.append([
                r["razon_social"], r["cuit"],
                l.get("corredor",""),
                l.get("linea",""), l.get("tramo",""),
                l.get("kmT",""), l.get("svcs",""),
                km_prod, l.get("ing",""),
            ])

    for col, w in enumerate([35,18,14,16,45,12,12,12,14], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    conn2.close()

    # Guardar y devolver
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    per = f"_{anio}_{str(mes).zfill(2)}" if anio and mes else ""
    path = BASE_DIR / f"export_ddjj{per}_{ts}.xlsx"
    wb.save(path)

    return FileResponse(
        path=str(path),
        filename=f"ERSeP_DDJJ{per}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── HEALTH CHECK ─────────────────────────────────────────────
@app.get("/health", include_in_schema=False)
def health():
    try:
        conn = get_db()
        total = conn.execute("SELECT COUNT(*) FROM ddjj").fetchone()[0]
        conn.close()
        return {"ok": True, "ddjj_registradas": total}
    except Exception:
        return {"ok": True, "ddjj_registradas": 0}


# ─── RUN LOCAL ────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
